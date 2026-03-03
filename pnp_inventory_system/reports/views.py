from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models import Count, Q
from django.core.paginator import Paginator
import csv
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from assets.models import Asset
from branches.models import Branch
from .models import Report
from users.views import role_required
from audit_logs.models import AuditLog

@login_required
def reports_home(request):
    """
    Main reports page
    """
    return render(request, 'reports/reports_home.html')

@login_required
def monthly_report(request):
    """
    Generate monthly inventory report
    """
    if request.method == 'POST':
        form_data = request.POST
        month = int(form_data.get('month'))
        year = int(form_data.get('year'))
        branch_id = form_data.get('branch')
        export_format = form_data.get('format')
        
        # Calculate date range
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        # Get assets based on user role and selected branch
        if request.user.role == 'provincial_admin':
            if branch_id:
                assets = Asset.objects.filter(branch_id=branch_id)
                branch = Branch.objects.get(id=branch_id)
            else:
                assets = Asset.objects.all()
                branch = None
        else:
            assets = Asset.objects.filter(branch=request.user.branch)
            branch = request.user.branch
        
        # Filter assets created or updated within the month
        monthly_assets = assets.filter(
            Q(created_at__date__range=(start_date, end_date)) |
            Q(updated_at__date__range=(start_date, end_date))
        ).select_related('branch')
        
        # Generate report
        if export_format == 'pdf':
            response = generate_pdf_report(monthly_assets, f"Monthly Report - {start_date.strftime('%B %Y')}", branch)
            # Log the report generation
            AuditLog.log_action(
                user=request.user,
                action='report',
                asset=None,
                description=f"Generated monthly report for {start_date.strftime('%B %Y')}" + (f" - {branch.name}" if branch else ""),
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            return response
        elif export_format == 'excel':
            response = generate_excel_report(monthly_assets, f"Monthly Report - {start_date.strftime('%B %Y')}", branch)
            # Log the report generation
            AuditLog.log_action(
                user=request.user,
                action='report',
                asset=None,
                description=f"Generated monthly Excel report for {start_date.strftime('%B %Y')}" + (f" - {branch.name}" if branch else ""),
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            return response
        elif export_format == 'csv':
            response = generate_csv_report(monthly_assets, f"Monthly Report - {start_date.strftime('%B %Y')}", branch)
            # Log the report generation
            AuditLog.log_action(
                user=request.user,
                action='report',
                asset=None,
                description=f"Generated monthly CSV report for {start_date.strftime('%B %Y')}" + (f" - {branch.name}" if branch else ""),
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            return response
    
    context = {
        'months': [(i, datetime(2024, i, 1).strftime('%B')) for i in range(1, 13)],
        'years': range(2020, timezone.now().year + 1),
        'current_month': timezone.now().month,
        'current_year': timezone.now().year,
    }
    
    # Add branches for Provincial Admin
    if request.user.role == 'provincial_admin':
        context['branches'] = Branch.objects.filter(is_active=True)
    
    return render(request, 'reports/monthly_report.html', context)

@login_required
def quarterly_report(request):
    """
    Generate quarterly inventory report
    """
    if request.method == 'POST':
        form_data = request.POST
        year = int(form_data.get('year'))
        quarter = int(form_data.get('quarter'))
        branch_id = form_data.get('branch')
        export_format = form_data.get('format')
        
        # Calculate date range for quarter
        quarter_start_month = (quarter - 1) * 3 + 1
        start_date = datetime(year, quarter_start_month, 1).date()
        end_date = datetime(year, quarter_start_month + 3, 1).date() - timedelta(days=1)
        
        # Get assets based on user role and selected branch
        if request.user.role == 'provincial_admin':
            if branch_id:
                assets = Asset.objects.filter(branch_id=branch_id)
                branch = Branch.objects.get(id=branch_id)
            else:
                assets = Asset.objects.all()
                branch = None
        else:
            assets = Asset.objects.filter(branch=request.user.branch)
            branch = request.user.branch
        
        # Filter assets created or updated within the quarter
        quarterly_assets = assets.filter(
            Q(created_at__date__range=(start_date, end_date)) |
            Q(updated_at__date__range=(start_date, end_date))
        ).select_related('branch')
        
        # Generate report
        if export_format == 'pdf':
            response = generate_pdf_report(quarterly_assets, f"Quarterly Report - Q{quarter} {year}", branch)
            # Log the report generation
            AuditLog.log_action(
                user=request.user,
                action='report',
                asset=None,
                description=f"Generated quarterly report for Q{quarter} {year}" + (f" - {branch.name}" if branch else ""),
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            return response
        elif export_format == 'excel':
            response = generate_excel_report(quarterly_assets, f"Quarterly Report - Q{quarter} {year}", branch)
            # Log the report generation
            AuditLog.log_action(
                user=request.user,
                action='report',
                asset=None,
                description=f"Generated quarterly Excel report for Q{quarter} {year}" + (f" - {branch.name}" if branch else ""),
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            return response
        elif export_format == 'csv':
            response = generate_csv_report(quarterly_assets, f"Quarterly Report - Q{quarter} {year}", branch)
            # Log the report generation
            AuditLog.log_action(
                user=request.user,
                action='report',
                asset=None,
                description=f"Generated quarterly CSV report for Q{quarter} {year}" + (f" - {branch.name}" if branch else ""),
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            return response
    
    context = {
        'quarters': [(1, 'Q1'), (2, 'Q2'), (3, 'Q3'), (4, 'Q4')],
        'years': range(2020, timezone.now().year + 1),
        'current_quarter': (timezone.now().month - 1) // 3 + 1,
        'current_year': timezone.now().year,
    }
    
    # Add branches for Provincial Admin
    if request.user.role == 'provincial_admin':
        context['branches'] = Branch.objects.filter(is_active=True)
    
    return render(request, 'reports/quarterly_report.html', context)

@login_required
def custom_report(request):
    """
    Generate custom date range report
    """
    if request.method == 'POST':
        form_data = request.POST
        start_date = datetime.strptime(form_data.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(form_data.get('end_date'), '%Y-%m-%d').date()
        branch_id = form_data.get('branch')
        export_format = form_data.get('format')
        
        # Get assets based on user role and selected branch
        if request.user.role == 'provincial_admin':
            if branch_id:
                assets = Asset.objects.filter(branch_id=branch_id)
                branch = Branch.objects.get(id=branch_id)
            else:
                assets = Asset.objects.all()
                branch = None
        else:
            assets = Asset.objects.filter(branch=request.user.branch)
            branch = request.user.branch
        
        # Filter assets created or updated within the date range
        custom_assets = assets.filter(
            Q(created_at__date__range=(start_date, end_date)) |
            Q(updated_at__date__range=(start_date, end_date))
        ).select_related('branch')
        
        # Generate report
        if export_format == 'pdf':
            return generate_pdf_report(custom_assets, f"Custom Report - {start_date} to {end_date}", branch)
        elif export_format == 'excel':
            return generate_excel_report(custom_assets, f"Custom Report - {start_date} to {end_date}", branch)
        elif export_format == 'csv':
            return generate_csv_report(custom_assets, f"Custom Report - {start_date} to {end_date}", branch)
    
    context = {}
    
    # Add branches for Provincial Admin
    if request.user.role == 'provincial_admin':
        context['branches'] = Branch.objects.filter(is_active=True)
    
    return render(request, 'reports/custom_report.html', context)

def generate_pdf_report(assets, title, branch=None):
    """
    Generate PDF report using ReportLab
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    elements.append(Paragraph(title, title_style))
    
    # Branch info (if specified)
    if branch:
        elements.append(Paragraph(f"Branch: {branch.name}", styles['Normal']))
        elements.append(Spacer(1, 12))
    
    # Report date
    elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary statistics
    total_assets = assets.count()
    active_assets = assets.filter(status='active').count()
    under_repair = assets.filter(status='under_repair').count()
    missing = assets.filter(status='missing').count()
    condemned = assets.filter(status='condemned').count()
    
    summary_data = [
        ['Summary Statistics', 'Count'],
        ['Total Assets', str(total_assets)],
        ['Active Assets', str(active_assets)],
        ['Under Repair', str(under_repair)],
        ['Missing', str(missing)],
        ['Condemned', str(condemned)],
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Asset details table
    if assets.exists():
        headers = ['Property Number', 'Type', 'Brand', 'Model', 'Status', 'Branch', 'Date Acquired']
        data = [headers]
        
        for asset in assets:
            data.append([
                asset.property_number,
                asset.get_asset_type_display(),
                asset.brand,
                asset.model,
                asset.get_status_display(),
                asset.branch.name if asset.branch else 'N/A',
                asset.date_acquired.strftime('%Y-%m-%d')
            ])
        
        asset_table = Table(data)
        asset_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(asset_table)
    
    doc.build(elements)
    return response

def generate_excel_report(assets, title, branch=None):
    """
    Generate Excel report using openpyxl
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Branch info
    row = 2
    if branch:
        ws[f'A{row}'] = f"Branch: {branch.name}"
        row += 1
    
    ws[f'A{row}'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    row += 2
    
    # Summary statistics
    ws[f'A{row}'] = "Summary Statistics"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    total_assets = assets.count()
    active_assets = assets.filter(status='active').count()
    under_repair = assets.filter(status='under_repair').count()
    missing = assets.filter(status='missing').count()
    condemned = assets.filter(status='condemned').count()
    
    summary_data = [
        ['Total Assets', total_assets],
        ['Active Assets', active_assets],
        ['Under Repair', under_repair],
        ['Missing', missing],
        ['Condemned', condemned],
    ]
    
    for label, count in summary_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = count
        row += 1
    
    row += 2
    
    # Asset details headers
    headers = ['Property Number', 'Type', 'Brand', 'Model', 'Serial Number', 'Processor', 'RAM', 'Storage', 
               'Status', 'Date Acquired', 'Warranty Expiry', 'Assigned Personnel', 'Branch']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row += 1
    
    # Asset data
    for asset in assets:
        data = [
            asset.property_number,
            asset.get_asset_type_display(),
            asset.brand,
            asset.model,
            asset.serial_number,
            asset.processor,
            asset.ram,
            asset.storage,
            asset.get_status_display(),
            asset.date_acquired.strftime('%Y-%m-%d'),
            asset.warranty_expiration.strftime('%Y-%m-%d') if asset.warranty_expiration else '',
            asset.assigned_personnel,
            asset.branch.name if asset.branch else 'N/A'
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response

def generate_csv_report(assets, title, branch=None):
    """
    Generate CSV report
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.csv"'
    
    writer = csv.writer(response)
    
    # Write header information
    writer.writerow([title])
    if branch:
        writer.writerow([f"Branch: {branch.name}"])
    writer.writerow([f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    writer.writerow([])
    
    # Write summary statistics
    writer.writerow(['Summary Statistics'])
    total_assets = assets.count()
    active_assets = assets.filter(status='active').count()
    under_repair = assets.filter(status='under_repair').count()
    missing = assets.filter(status='missing').count()
    condemned = assets.filter(status='condemned').count()
    
    writer.writerow(['Total Assets', total_assets])
    writer.writerow(['Active Assets', active_assets])
    writer.writerow(['Under Repair', under_repair])
    writer.writerow(['Missing', missing])
    writer.writerow(['Condemned', condemned])
    writer.writerow([])
    
    # Write asset data
    headers = ['Property Number', 'Type', 'Brand', 'Model', 'Serial Number', 'Processor', 'RAM', 'Storage', 
               'Status', 'Date Acquired', 'Warranty Expiry', 'Assigned Personnel', 'Branch']
    writer.writerow(headers)
    
    for asset in assets:
        writer.writerow([
            asset.property_number,
            asset.get_asset_type_display(),
            asset.brand,
            asset.model,
            asset.serial_number,
            asset.processor,
            asset.ram,
            asset.storage,
            asset.get_status_display(),
            asset.date_acquired.strftime('%Y-%m-%d'),
            asset.warranty_expiration.strftime('%Y-%m-%d') if asset.warranty_expiration else '',
            asset.assigned_personnel,
            asset.branch.name if asset.branch else 'N/A'
        ])
    
    return response

@login_required
def report_history(request):
    """
    View report generation history
    """
    if request.user.role == 'provincial_admin':
        reports = Report.objects.all().order_by('-created_at')
    else:
        reports = Report.objects.filter(branch=request.user.branch).order_by('-created_at')
    
    paginator = Paginator(reports, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'reports/report_history.html', {'page_obj': page_obj})
